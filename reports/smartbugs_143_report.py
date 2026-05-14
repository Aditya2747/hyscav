"""
SmartBugs 143 Contracts Report Generator - Enhanced Version.
Generates comprehensive Excel reports with SWC IDs, severity, descriptions.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# SmartBugs dataset statistics - 143 contracts in 10 categories
CATEGORY_STATS = {
    'access_control': {'count': 18, 'label': 1, 'vuln': 'Access Control', 'swc': 'SWC-100', 'severity': 'HIGH', 'description': 'Vulnerabilities related to access control'},
    'arithmetic': {'count': 15, 'label': 2, 'vuln': 'Integer Overflow', 'swc': 'SWC-101', 'severity': 'HIGH', 'description': 'Arithmetic bugs including overflow/underflow'},
    'bad_randomness': {'count': 8, 'label': 1, 'vuln': 'Weak Randomness', 'swc': 'SWC-120', 'severity': 'MEDIUM', 'description': 'Insecure source of randomness'},
    'denial_of_service': {'count': 6, 'label': 1, 'vuln': 'DOS', 'swc': 'SWC-110', 'severity': 'MEDIUM', 'description': 'Denial of service vulnerabilities'},
    'front_running': {'count': 4, 'label': 2, 'vuln': 'Front Running', 'swc': 'SWC-132', 'severity': 'HIGH', 'description': 'Transaction ordering conflicts'},
    'other': {'count': 3, 'label': 0, 'vuln': 'Various', 'swc': 'Various', 'severity': 'LOW', 'description': 'Miscellaneous vulnerabilities'},
    'reentrancy': {'count': 31, 'label': 2, 'vuln': 'Reentrancy', 'swc': 'SWC-107', 'severity': 'HIGH', 'description': 'Reentrancy attack vulnerability'},
    'short_addresses': {'count': 1, 'label': 1, 'vuln': 'Short Address', 'swc': 'SWC-133', 'severity': 'MEDIUM', 'description': 'Short address attack'},
    'time_manipulation': {'count': 5, 'label': 1, 'vuln': 'Timestamp', 'swc': 'SWC-116', 'severity': 'MEDIUM', 'description': 'Block timestamp dependency'},
    'unchecked_low_level_calls': {'count': 52, 'label': 2, 'vuln': 'Unchecked Call', 'swc': 'SWC-104', 'severity': 'HIGH', 'description': 'Unchecked low-level call return values'},
}

RISK_LABELS = {0: 'LOW', 1: 'MEDIUM', 2: 'HIGH'}


def create_summary_report():
    """Create full SmartBugs 143 summary with all parameters."""
    data = []
    contract_num = 1
    
    for category, stats in CATEGORY_STATS.items():
        count = stats['count']
        label = stats['label']
        
        for i in range(count):
            data.append({
                'ID': contract_num,
                'Category': category.replace('_', ' ').title(),
                'Contract Index': i + 1,
                'Risk Level': RISK_LABELS[label],
                'Severity': stats.get('severity', 'MEDIUM'),
                'SWC ID': stats.get('swc', 'N/A'),
                'Vulnerability Type': stats['vuln'],
                'Description': stats.get('description', ''),
                'Analysis Method': 'Category-based Heuristic',
                'Dataset': 'SmartBugs',
            })
            contract_num += 1
    
    return pd.DataFrame(data)


def create_category_stats():
    """Create category statistics with all parameters."""
    data = []
    
    for category, stats in CATEGORY_STATS.items():
        data.append({
            'Category': category.replace('_', ' ').title(),
            'Contract Count': stats['count'],
            'Risk Level': RISK_LABELS[stats['label']],
            'Severity': stats.get('severity', 'MEDIUM'),
            'SWC ID': stats.get('swc', 'N/A'),
            'Vulnerability Type': stats['vuln'],
            'Description': stats.get('description', ''),
            'Share (%)': round((stats['count'] / 143) * 100, 1),
        })
    
    df = pd.DataFrame(data)
    return df.sort_values('Severity', ascending=False)


def apply_styling(ws, df):
    """Apply styling to Excel worksheet."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    risk_fills = {
        'HIGH': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'MEDIUM': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        'LOW': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }
    
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Rows
    for row_idx in range(2, ws.max_row + 1):
        risk_cell = ws.cell(row=row_idx, column=4)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        risk_val = risk_cell.value
        if risk_val in risk_fills:
            risk_cell.fill = risk_fills[risk_val]
            risk_cell.font = Font(bold=True)


def save_excel(df, filepath, sheet_name):
    """Save DataFrame to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    apply_styling(ws, df)
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)
    
    wb.save(filepath)
    print(f"  Saved: {filepath}")


def main():
    """Generate SmartBugs 143 Report."""
    print("=" * 60)
    print("SmartBugs 143 Contracts Analysis Report")
    print("=" * 60)
    
    os.makedirs('reports', exist_ok=True)
    
    # 1. Category Stats
    print("\n[1] Creating category statistics...")
    cat_df = create_category_stats()
    total = cat_df['Contract Count'].sum()
    
    print(f"  Total contracts: {total}")
    print(f"  HIGH severity: {cat_df[cat_df['Severity']=='HIGH']['Contract Count'].sum()}")
    print(f"  MEDIUM severity: {cat_df[cat_df['Severity']=='MEDIUM']['Contract Count'].sum()}")
    print(f"  LOW severity: {cat_df[cat_df['Severity']=='LOW']['Contract Count'].sum()}")
    
    # 2. Full List
    print("\n[2] Creating full contract list...")
    full_df = create_summary_report()
    print(f"  Total: {len(full_df)} contracts")
    
    # 3. Save Reports
    print("\n[3] Saving Excel reports...")
save_excel(cat_df, 'reports/smartbugs_143_categories_v2.xlsx', 'Categories')
    save_excel(full_df, 'reports/smartbugs_143_all_v2.xlsx', 'All Contracts')
    
    # 4. CSV  
    full_df.to_csv('reports/smartbugs_143_all_v2.csv', index=False)
    print(f"  Saved: reports/smartbugs_143_all_v2.csv")
    
    print("\n" + "=" * 60)
    print("Report Generation Complete!")
    print("=" * 60)
    
    return full_df


if __name__ == "__main__":
    main()
